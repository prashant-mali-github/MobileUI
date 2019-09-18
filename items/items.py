from io import BytesIO
import io
import img2pdf
import os
from flask import Blueprint, render_template, redirect, request,flash,url_for,session,g,send_from_directory,send_file
from models.create_models import Items, Bill, SalesItems, Customer
from orders.orders import order
from openpyxl import load_workbook
import base64
import functools
import random
from app import db
from io import BytesIO

item = Blueprint('item', __name__, template_folder='template')


@item.route('/addoneitem/', methods=['GET', 'POST'])
def addoneitems():
    if request.method == "POST":
        error=None
        item_name = request.form.get('iname')
        item_quantity = request.form.get('quantity')
        item_price = request.form['price']
        import pyqrcode
        url = pyqrcode.create(f"{item_name}---------{item_price}")

        path = 'E:\\DeployMobileUI\\items\\barcode'
        image_name = path + '\\' + item_name + '.png'
        os.chdir(path)
        url.png(image_name, scale=8)
        with open(image_name,"rb+") as f:
            x = f.read()
        if not item_name or not item_quantity or not item_price:
            flash('Please enter all the fields', 'error')
            return redirect(url_for('item.additems'))
        else:
            item = Items(item_name=item_name, item_quantity=item_quantity, item_price=item_price, barcode=x)
            db.session.add(item)
            db.session.commit()
            return redirect(url_for('item.viewitems'))
    return render_template("static priya mobile/items/advance_insert.html")


@item.route('/additem/', methods=['GET', 'POST'])
def additems():
    if request.method == "POST":
        error=None
        item_name = request.form.getlist('iname')
        item_quantity = request.form.get('quantity')
        item_price = request.form['price']
        print(item_name,".....itemlist")
        import pyqrcode
        url = pyqrcode.create(f"{item_name}---------{item_price}")
        # path = 'E:\\DeployMobileUI\\items\\barcode'
        # image_name = path + '\\' + item_name + '.png'
        # os.chdir(path)
        # url.png(image_name, scale=8)
        # with open(image_name,"rb+") as f:
        #     x = f.read()
        if not item_name or not item_quantity or not item_price:
            flash('Please enter all the fields', 'error')
            return redirect(url_for('item.additems'))
        else:
            if len(item_name) == 1:
                path = 'E:\\DeployMobileUI\\items\\barcode'
                image_name = path + '\\' + item_name + '.png'
                url = pyqrcode.create(f"{image_name}---------\n{item_price}")
                os.chdir(path)
                url.png(image_name, scale=8)
                with open(image_name, "rb+") as f:
                    x = f.read()
                item = Items(item_name=item_name, item_quantity=item_quantity, item_price=item_price, barcode=x)
                db.session.add(item)
                db.session.commit()
                return redirect(url_for('item.viewitems'))
            elif len(item_name) > 1:
                for item_n in item_name:
                    path = 'E:\\DeployMobileUI\\items\\barcode'
                    item_path = path + '\\' + item_n + '.png'
                    os.chdir(path)
                    url = pyqrcode.create(f"{item_n}---------\n{item_price}")
                    url.png(item_path, scale=8)
                    with open(item_path, "rb+") as f:
                        x = f.read()
                    item = Items(item_name=item_n, item_quantity=item_quantity, item_price=item_price, barcode=x)
                    db.session.add(item)
                    db.session.commit()
                return redirect(url_for('item.viewitems'))
    return render_template("static priya mobile/items/insert_item.html")


@item.route('/viewitems')
def viewitems():
    items = Items.query.all()
    # path = 'E:\\DeployMobileUI\\items\\barcode'
    # for i in items:
    #     # with open(path + '\\' + i.item_name + '.png', "wb") as f:
    #     #     f.write(bytearray(i.barcode))
    #     #     f.close()
    #     data = i.barcode
    #     print(data)
    #     os.chdir(path)
    #     bytes = i.barcode
    #     image = Image.open(io.BytesIO(bytes))
    #     image.save(path)
    return render_template('static priya mobile/items/viewitems.html', items=items)


@item.route('/download_qr/<id>')
def download_qr(id):
    item = Items.query.get(int(id))
    barcode_data = item.barcode
    file_name = str(item.item_name) + '.png'
    path = 'E:\\DeployMobileUI\\items\\download_qr'
    os.chdir(path)
    with open(file_name, "wb+") as file:
        file.write(barcode_data)
    return send_file(BytesIO(barcode_data), attachment_filename=file_name, as_attachment=True)


@item.route('/all_barcode')
def all_barcode():
    imagelist = []
    imagelist_obj = Items.query.all()
    path = 'E:\\DeployMobileUI\\items\\download_qr'
    for image in imagelist_obj:
        file_name = str(image.item_name) + '.png'
        barcode_data = image.barcode
        os.chdir(path)
        with open(file_name, "wb+") as file:
            file.write(barcode_data)
    with open("all_item_barcode.pdf", "wb") as f:
        f.write(img2pdf.convert([i for i in os.listdir('E:\\DeployMobileUI\\items\\download_qr') if i.endswith(".png")]))
    return redirect(url_for('item.viewitems'))


def get_list_of_file_names():
    list_of_filename = []
    for root, dirs, files in os.walk("download_qr"):
        for filename in files:
            list_of_filename.append(filename)
    return list_of_filename

@item.route('/reports')
def report():
    items = Items.query.all()
    import xlsxwriter
    # Workbook() takes one, non-optional, argument
    # which is the filename that we want to create.
    workbook = xlsxwriter.Workbook('allitems.xlsx')

    # The workbook object is then used to add new
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()
    row = 0
    column = 0
        # # Use the worksheet object to write
        # # data via the write() method.
        # worksheet.write('A1', i.i_id)
        # worksheet.write('B1', i.item_name)
        # worksheet.write('C1', i.item_quantity)
        # worksheet.write('D1', i.item_price)
    # Finally, close the Excel file
    for item in items:
        worksheet.write(row, column, item.i_id)
        worksheet.write(row, column+1, item.item_name)
        worksheet.write(row, column+2, item.item_quantity)
        worksheet.write(row, column+3, item.item_price)
        worksheet.write(row, column+4, item.purchase_date)
        # incrementing the value of row by one
        # with each iteratons.
        row += 1
    workbook.close()

    file_name = 'document_template.xltx'
    wb = load_workbook('hello.xlsx')
    wb.save(file_name)
    return send_from_directory(file_name, as_attachment=True)
    # via the close() method.


@item.route('/<int:id>/update', methods=('GET', 'POST'))
def updateitem(id):
    error = None
    item = Items.query.filter_by(i_id=int(id)).first()
    #
    # print("................post",post.title)
    if request.method == 'POST':

        item_name = request.form['iname']
        item_quantity = request.form['quantity']
        item_price = request.form['price']
        error = None

        if not item_name or not item_quantity or not item_price:
            flash(error)
        if error is not None:
            flash(error)
        else:
            item.item_name = item_name
            item.item_quantity = item_quantity
            item.item_price = item_price
            db.session.add(item)
            db.session.commit()
            return redirect(url_for('item.viewitems'))
    return render_template('static priya mobile/items/updateitem.html', item=item)



@item.route('/<int:id>/saleitem', methods=['GET', 'POST'])
def saleitem(id):
    customers = Customer.query.all()
    if request.method == "POST":
        error=None
        c_name = request.form.get('c_id')
        quantity = request.form['quantity']
        if not c_name or not quantity:
            flash('Please enter all the fields', 'error')
        else:
            customer = Customer.query.filter_by(customer_name=c_name).first()
            item = Items.query.filter_by(i_id=id).first()
            if not customer:
                error = "Invalid customer ID"
                flash(error)
            if not item:
                error = "Invalid Item ID"
                flash(error)
            available_quantity = int(item.item_quantity)
            sale_quantity = int(quantity)
            bill_amount = int(item.item_price) * int(sale_quantity)
            order_id = random.randint(1000000000,999999999999)
            import pyqrcode
            url = pyqrcode.create(f"{item.item_name}------\n---{item.item_price}--\n----{item.item_quantity}---\n--Total={bill_amount}")
            url.png('orders/bill.png', scale=8)
            with open("orders/bill.png", "rb+") as f:
                x = f.read()
            if available_quantity >= sale_quantity:
                order = SalesItems(id=order_id, c_id=customer.c_id, i_id=id, sale_quantity=sale_quantity)
                db.session.add(order)
                db.session.commit()
                item.item_quantity = available_quantity - sale_quantity
                bill = Bill(o_id=order_id, bill_amount=bill_amount, bill_barcode=x)
                db.session.add_all([item, bill])
                db.session.commit()
                return redirect(url_for('order.totalsale'))
            else:
                error="Please enter less sale quantity than available quantity"
                flash(error)

    return render_template("static priya mobile/items/saleitem.html",customers=customers)

@item.route('/<int:id>/deleteorder', methods=('GET', 'POST'))
def deleteorders(id):
    error=None
    order = SalesItems.query.filter_by(id=int(id)).first()
    item = Items.query.filter_by(i_id=int(order.i_id)).first()
    available_quantity = int(item.item_quantity)
    sale_quantity = int(order.sale_quantity)
    item.item_quantity = available_quantity + sale_quantity
    bill = Bill.query.filter_by(o_id=int(order.id)).first()
    db.session.add(item)
    db.session.commit()
    db.session.delete(order)
    db.session.delete(bill)
    db.session.commit()
    return redirect(url_for('order.vieworders'))



@item.route('/<int:id>/delete', methods=('GET', 'POST'))
def deleteitem(id):
    error = None
    item = Items.query.filter_by(i_id=int(id)).first()
    db.session.delete(item)
    db.session.commit()
    return redirect(url_for('item.viewitems'))


@item.route('/searchitems')
def searchitems():
    items = Items.query.all()
    return render_template('static priya mobile/base.html',items=items)


@item.route('/search_item')
def search_items():
    items = Items.query.all()
    return render_template('static priya mobile/items/search_items.html', items=items)
